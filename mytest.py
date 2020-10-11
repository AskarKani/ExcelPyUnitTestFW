import os
import subprocess
import cppyy
import shutil
import openpyxl
import sys
import math
from pathlib import Path

class Mytest:
    def __init__(self, excel_path):
        self.unit_test_out = Path(os.getcwd()) / "unit_test_out"
        if not os.path.isdir(self.unit_test_out):
            os.mkdir(self.unit_test_out)
        self.my_cppyy = cppyy
        self.excel_path = excel_path
        self.flags = ""

    def terminal_exec(self, cmd):
        try:
            outstr = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, check=True)
            logdata = outstr.stdout.decode()
            return logdata
        except subprocess.CalledProcessError as e:
            logdata = "\n\nError:\n" + e.stdout.decode('utf-8')
            return logdata

    def split_check_path(self, header, entries, is_path_check=True ,is_folder=False):
        # Check for brackets "[ ]"
        if entries.find("[") == -1 or entries.find("]") == -1:
            sys.exit(f"!! EXITING !! {header} value is missing \"[\" or \"]\" ")

        # Check for empty entries
        paths_check_empty = entries.strip("[]").strip(" ")
        if not paths_check_empty:
            print(f"{header} has no entries")
            return list(paths_check_empty)

        # Split the paths and convert to list
        entries = entries.strip('[]').split(',')
        for i in range(0, len(entries)):
            entries[i] = entries[i].lstrip(" ").rstrip(" ")
        #Check if file/folder path is invalid
        if is_path_check:
            for entry in entries:
                if not is_folder:
                    if not os.path.isfile(entry):
                        sys.exit(f"!! EXITING !! {header}:{entry} path is incorrect...")
                else:
                    if not os.path.isdir(entries[i]):
                        sys.exit(f"!! EXITING !! {header}:{entry} folder path is incorrect...")
        return entries

    def read_excel(self):
        self.book = openpyxl.load_workbook(self.excel_path, read_only=False, data_only=True)
        self.sheet = self.book.active
        self.lib_include_path_s = self.split_check_path("$LIB_INC_PATH(s)", self.sheet['B1'].value, is_folder=True)
        self.lib_name_s = self.split_check_path("$LIB_NAME(s)", self.sheet['B2'].value, is_path_check=False)
        self.include_file_s = self.split_check_path("$INC_FILE", self.sheet['B3'].value)
        self.func_file_s = self.split_check_path("$FUNC_FILE", self.sheet['B4'].value)
        self.include_path_gcc_s = self.split_check_path("$INC_PATH_GCC(s)", self.sheet['B5'].value, is_folder=True)
        self.flags_gcc = self.split_check_path("$FLAGS_GCC", self.sheet['B6'].value, is_path_check=False)
        self.func_name = self.sheet['B7'].value
        if not self.func_name:
            print(f"enter func_name in {self.excel_path}")
            exit(1)

    def append_option(self, option, paths):
        append_string = ""
        for path in paths:
            append_string = f"{append_string} {option} {path}"
        return append_string

    def compile(self):
        '''
         # g++ -fpic -c add.cpp
         # g++ -shared -o libadd.so add.o
        '''
        lib_include_paths = self.append_option("-L", self.lib_include_path_s)
        include_path_gcc = self.append_option("-I", self.include_path_gcc_s)
        lib_names = self.append_option("-l", self.lib_name_s)
        flags = self.append_option("", self.flags_gcc)

        out_file = []
        for file in self.func_file_s:
            file_cpp_name = os.path.splitext(os.path.basename(file))[0]
            out_cpp_path = self.unit_test_out / file_cpp_name
            self.terminal_exec(f"g++ -fpic -c {file} {include_path_gcc} -o {out_cpp_path} {flags}")
            out_file.append(out_cpp_path)

        lib_name = "libtest.so"
        lib_out_path = self.unit_test_out / lib_name
        self.terminal_exec(f"g++ -shared -o {lib_out_path} {self.append_option('', out_file)} {lib_include_paths} {lib_names} ")

        for file in self.include_file_s:
            self.my_cppyy.include(file)
        # for path in self.lib_include_path_s:
        #     print(os.path.abspath(path))
        #     os.environ['LD_LIBRARY_PATH'] = str(os.path.abspath(path))
        #     self.my_cppyy.add_library_path("/home/askar/my_projects/cpp/unit_test_cpp/examples/lib_example/")
        try:
            self.my_cppyy.load_library(str(lib_out_path))
        except RuntimeError:
            print("ERROR : Library path is missing or error in compilation..")
            exit(1)

    def clean(self):
        if os.path.isdir(self.unit_test_out):
            shutil.rmtree(self.unit_test_out)

    def test(self):
        self.read_excel()
        self.compile()
        print(f"Executing unit test for function : {self.func_name}")
        column = self.sheet.max_column
        max_row = self.sheet.max_row
        expected_result_col = 0
        header_row = 8

        for i in range(1, column + 1):
            cell_value = self.sheet.cell(row=header_row, column=i).value
            if cell_value == "Expected Result":
                expected_result_col = i
        result_obtained_col = expected_result_col + 1
        result_col = expected_result_col + 2
        test_case_start_row = 9

        pass_count = fail_count = 0
        greenFill = openpyxl.styles.PatternFill(start_color='FF00FF00',
                                                end_color='FF00FF00',
                                                fill_type='solid')
        redFill = openpyxl.styles.PatternFill(start_color='FFFF0000',
                                              end_color='FFFF0000',
                                              fill_type='solid')
        for j in range(test_case_start_row, max_row + 1):
            test_params = []
            for i in range(2, expected_result_col):
                test_params.append(self.sheet.cell(row=j, column=i).value)
            expected_result = self.sheet.cell(row=j, column=expected_result_col).value

            temp_func = f"locals()['temp']=self.my_cppyy.gbl.{self.func_name}("
            for i in test_params:
                try:
                    int(i)
                    temp_func += str(i) + ","
                except ValueError:
                    temp_func += "\"" + str(i) + "\"" + ","
            func = temp_func + ")"
            exec(func)
            result_obtained = locals()['temp']
            self.sheet.cell(row=j, column=result_obtained_col).value = result_obtained

            try:
                int(expected_result)
                # FIXME: float value representation of python is different in precision.So comparision is done using isclose
                if math.isclose(expected_result, result_obtained, abs_tol=0.05):
                    result = "PASS"
                    pass_count = pass_count + 1
                    self.sheet.cell(row=j, column=result_col).fill = greenFill
                else:
                    result = "FAIL"
                    fail_count = fail_count + 1
                    self.sheet.cell(row=j, column=result_col).fill = redFill
            except ValueError:
                if expected_result == result_obtained:
                    result = "PASS"
                    pass_count = pass_count + 1
                    self.sheet.cell(row=j, column=result_col).fill = greenFill
                else:
                    result = "FAIL"
                    fail_count = fail_count + 1
                    self.sheet.cell(row=j, column=result_col).fill = redFill
            self.sheet.cell(row=j, column=result_col).value = result

        self.book.save(self.excel_path)
        print(f"Total test cases executed : {max_row - test_case_start_row + 1}")
        print(f"Total PASS : {pass_count}")
        print(f"Total FAIL : {fail_count}")
        print()
        self.clean()
        